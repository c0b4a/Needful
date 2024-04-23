#!/usr/bin/env python
import sys
import pandas
import openpyxl
from openpyxl import Workbook


def openWorkbook(link) -> Workbook:
    wb = Workbook(link)
    #assigns active worksheet based on last used
    global ws
    ws = wb.active
    #delete me
    print(wb.sheetnames)
    return wb

def closeWorkbook(link, wb) -> None:
    wb.save(link)

#def readWorksheet(ws):
#    for row in ws.values:
#        for value in row:
#            print(value)

def main() -> None:
    link = sys.argv[1]
    wb = openWorkbook(link)
    #readWorksheet(ws)
    closeWorkbook(link, wb)
    sys.exit()

if __name__ == '__main__':
    main()